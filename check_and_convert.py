#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""
Lee el JSON completo y convierte a Excel con todos los registros
"""
import json
import os
from datetime import datetime

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
except ImportError:
    print("openpyxl no está instalado.")
    exit(1)

RESULTS_DIR = 'resultados'
JSON_FILE = os.path.join(RESULTS_DIR, 'extraction_20251112_122412.json')

print(f"Leyendo {JSON_FILE}...")
try:
    with open(JSON_FILE, 'r', encoding='utf-8') as f:
        data = json.load(f)
except Exception as e:
    print(f"Error leyendo JSON: {e}")
    exit(1)

total_records = len(data)
print(f"✓ Total de registros encontrados: {total_records}")

if total_records == 0:
    print("No hay registros para convertir.")
    exit(1)

# Crear workbook
wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Inmobiliarias"

# Encabezados
headers = ['#', 'URL', 'Nombre Inmobiliaria', 'Cantidad de Inmuebles']
ws.append(headers)

# Aplicar estilos al encabezado
header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
header_font = Font(bold=True, color="FFFFFF")

for cell in ws[1]:
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

# Agregar datos
print("Agregando datos al Excel...")
for row_idx, record in enumerate(data, 2):
    ws[f'A{row_idx}'] = row_idx - 1  # Número de fila (1-based)
    ws[f'B{row_idx}'] = record.get('url', '')
    ws[f'C{row_idx}'] = record.get('titulo', '')
    ws[f'D{row_idx}'] = record.get('cantidad_inmuebles', 0)
    
    # Alinear
    for col in ['A', 'B', 'C', 'D']:
        ws[f'{col}{row_idx}'].alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)

# Ajustar ancho de columnas
ws.column_dimensions['A'].width = 6
ws.column_dimensions['B'].width = 70
ws.column_dimensions['C'].width = 40
ws.column_dimensions['D'].width = 25

# Guardar
excel_file = os.path.join(RESULTS_DIR, 'extraction_completo.xlsx')
wb.save(excel_file)

print(f"✓ Archivo Excel guardado en: {excel_file}")
print(f"✓ Total de filas: {total_records + 1} (incluye encabezado)")
print(f"✓ Archivo listo para descargar!")
