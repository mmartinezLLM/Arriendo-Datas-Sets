#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""
Convierte el JSON de resultados de extracción a un archivo Excel
"""
import json
import os
from datetime import datetime

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
except ImportError:
    print("openpyxl no está instalado. Instálalo con: pip install openpyxl")
    exit(1)

RESULTS_DIR = 'resultados'
JSON_FILE = os.path.join(RESULTS_DIR, 'extraction_20251112_122412.json')

# Leer JSON
print(f"Leyendo {JSON_FILE}...")
with open(JSON_FILE, 'r', encoding='utf-8') as f:
    data = json.load(f)

print(f"Total de registros: {len(data)}")

# Crear workbook
wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Inmobiliarias"

# Encabezados
headers = ['URL', 'Nombre Inmobiliaria', 'Cantidad de Inmuebles']
ws.append(headers)

# Aplicar estilos al encabezado
header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
header_font = Font(bold=True, color="FFFFFF")

for cell in ws[1]:
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

# Agregar datos
for row_idx, record in enumerate(data, 2):
    ws[f'A{row_idx}'] = record.get('url', '')
    ws[f'B{row_idx}'] = record.get('titulo', '')
    ws[f'C{row_idx}'] = record.get('cantidad_inmuebles', 0)
    
    # Alinear
    for col in ['A', 'B', 'C']:
        ws[f'{col}{row_idx}'].alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)

# Ajustar ancho de columnas
ws.column_dimensions['A'].width = 60
ws.column_dimensions['B'].width = 40
ws.column_dimensions['C'].width = 25

# Guardar
timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
excel_file = os.path.join(RESULTS_DIR, f'extraction_{timestamp}.xlsx')
wb.save(excel_file)

print(f"✓ Archivo Excel guardado en: {excel_file}")
print(f"✓ Total de filas: {len(data) + 1} (incluye encabezado)")
