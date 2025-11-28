#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""
Fusiona todos los archivos JSON de extracción (incluidos parciales) 
y elimina duplicados manteniendo el orden.
"""
import json
import os
from datetime import datetime
from collections import OrderedDict

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
except ImportError:
    print("openpyxl no está instalado.")
    exit(1)

RESULTS_DIR = 'resultados'

# Lista de archivos JSON a fusionar (en orden de preferencia)
json_files = [
    os.path.join(RESULTS_DIR, 'extraction_20251112_122412.json'),  # Final - preferencia
    os.path.join(RESULTS_DIR, 'extraction_partial_latest.json'),
    os.path.join(RESULTS_DIR, 'extraction_partial_20251112_121927.json'),
    os.path.join(RESULTS_DIR, 'extraction_partial_20251112_121400.json'),
]

all_data = {}  # Usar dict para evitar duplicados (clave: URL)

print("Leyendo y fusionando archivos JSON...")
for json_file in json_files:
    if os.path.exists(json_file):
        try:
            with open(json_file, 'r', encoding='utf-8') as f:
                data = json.load(f)
                if isinstance(data, list):
                    for record in data:
                        url = record.get('url', '')
                        if url:  # Solo si tiene URL
                            all_data[url] = record  # Sobrescribe si existe (toma la más reciente)
            print(f"  ✓ {os.path.basename(json_file)}: {len(data)} registros")
        except Exception as e:
            print(f"  ⚠️ Error leyendo {os.path.basename(json_file)}: {e}")
    else:
        print(f"  ⚠️ Archivo no encontrado: {os.path.basename(json_file)}")

# Convertir dict a lista manteniendo el orden de URLs únicas
unique_data = list(all_data.values())
total_records = len(unique_data)

print(f"\n✓ Total de registros únicos después de fusionar: {total_records}")

if total_records == 0:
    print("No hay registros para convertir.")
    exit(1)

# Crear workbook
wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Inmobiliarias"

# Encabezados
headers = ['#', 'URL', 'Nombre Inmobiliaria', 'Cantidad de Inmuebles']
ws.append(headers)

# Aplicar estilos al encabezado
header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
header_font = Font(bold=True, color="FFFFFF")

for cell in ws[1]:
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

# Agregar datos
print(f"Agregando {total_records} registros al Excel...")
for row_idx, record in enumerate(unique_data, 2):
    ws[f'A{row_idx}'] = row_idx - 1  # Número de fila (1-based)
    ws[f'B{row_idx}'] = record.get('url', '')
    ws[f'C{row_idx}'] = record.get('titulo', '')
    ws[f'D{row_idx}'] = record.get('cantidad_inmuebles', 0)
    
    # Alinear
    for col in ['A', 'B', 'C', 'D']:
        ws[f'{col}{row_idx}'].alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)

# Ajustar ancho de columnas
ws.column_dimensions['A'].width = 6
ws.column_dimensions['B'].width = 70
ws.column_dimensions['C'].width = 40
ws.column_dimensions['D'].width = 25

# Guardar
excel_file = os.path.join(RESULTS_DIR, 'extraction_completo_fusionado.xlsx')
wb.save(excel_file)

print(f"\n✓ Archivo Excel guardado en: {excel_file}")
print(f"✓ Total de filas: {total_records + 1} (incluye encabezado)")
print(f"✓ Archivo con TODOS los registros listo para descargar!")
